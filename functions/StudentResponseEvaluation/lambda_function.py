"""
Purpose:
This Lambda function reads student responses from an Excel file stored in S3, analyzes the responses using AWS Bedrock, and generates attendent status, scores and insights for each student.

Input:
- Excel file containing student responses

Output:
- Attendance status
- Scores and insights for each student response
"""
import json
import boto3
import openpyxl
from io import BytesIO

student_list = [
    112502020,
    113502002,
    113502004,
    113502005,
    113502006,
    113502008,
    113502009,
    113502010,
    113502011,
    113502012,
    113502013,
    113502015,
    113502016,
    113502017,
    113502019,
    113502021,
    113502022,
    113502024,
    113502025,
    113502026,
    113502027,
    113502028,
    113502501,
    113502502,
    113502503,
    113502504,
    113502505,
    113502506,
    113502507,
    113502508,
    113502509,
    113502510,
    113502511,
    113502512,
    113502513,
    113502514,
    113502515,
    113502517,
    113502519,
    113502520,
    113502521,
    113502522,
    113502523,
    113502524,
    113502525,
    113502526,
    113502527,
    113502528,
    113502529,
    113502530,
    113502531,
    113502532,
    113502533,
    113502534,
    113502535,
    113502536,
    113502537,
    113502538,
    113502540,
    113502541,
    113502542,
    113502543,
    113502545,
    113502546,
    113502547,
    113502548,
    113502549,
    113502551,
    113502552,
    113502553,
    113502554,
    113502555,
    113502557,
    113502559,
    113502560,
    113502566,
    113502568,
    113502569,
    113502570,
    113502571,
    113502572,
    113502573,
    113502574,
    113502576
]


# Initialize AWS clients
s3 = boto3.client('s3')
bedrock = boto3.client('bedrock-runtime', region_name="us-east-1")


def analyze_responses(question, responses, start_index=0):
    """Use AWS Bedrock to analyze responses and generate scores and insights."""
    modelId = 'anthropic.claude-3-haiku-20240307-v1:0'
    accept = 'application/json'
    contentType = 'application/json'

    rate_result = []
    
    batch_size = 7

    for i in range(start_index, len(responses.items()), batch_size):
        print(f"start from {i} to {i+batch_size}")
        # 獲取當前批次的學生 ID 和回應
        batch = {stu_id: responses[stu_id] for stu_id in list(responses.keys())[i:i + batch_size]}

        
        for stu_id, response in batch.items():
            # Prepare the request body for scoring
            batch_responses = "\n".join([f"學生ID: {stu_id}, 回覆: {response}" for stu_id, response in batch.items()])

        scoring_body = json.dumps({
            "anthropic_version": "bedrock-2023-05-31",
            "max_tokens": 400,
            "messages": [
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "text",
                            "text": f"""
                            題目: {question}  
                            學生回答: {batch_responses}
                            
                            請按照以下JSON格式回覆，不包含換行、空白，若有下一筆學生資料，請於大括號之間加入逗號區隔： 
                            {{"student_id": 學生ID,"score": 評分,"reason": 評分理由}},

                            規則：
                                1. 根據學生回答的努力程度和正確性打分（0-10）。
                                2. 若有文不對題、回應不完整、攻擊性言詞、答題態度輕浮，則斟酌扣分。
                                3. 評分理由請言簡意賅，在20字元內。

                            """
                        }
                    ]
                }
            ]
        })
        
        # Invoke the model for scoring
        scoring_response = bedrock.invoke_model(body=scoring_body, modelId=modelId, accept=accept, contentType=contentType)
        scoring_response_body = json.loads(scoring_response.get('body').read())

        
        clean_text = scoring_response_body["content"][0]["text"].replace("\n", "").replace("'", "\"") # string format
        try:
            if clean_text.startswith("{") and clean_text.endswith("}"):
                clean_text = clean_text.replace('},','}},')
                clean_lst = clean_text.split('},')
                # turn str to dict
                rate_result_list = [json.loads(i) for i in clean_lst]
                rate_result.extend(rate_result_list) 
            else:
                raise ValueError("Cleaned text is not a valid JSON object.")
            
        except Exception as e:
            print(f"Error encountered: {e}\nthe text is : {clean_lst}")
            print(f"Retrying Bedrock request from {i}...")
            return start_index

        
    print("Rate Results: ", rate_result)    
    return rate_result # [json1, json2, ...]

def calculate_attendance(sheet, attendance_column_index):
    """Calculate attendance based on the specified column index."""
    present_students = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[attendance_column_index]:  # Check if attendance is marked (assuming non-empty means present)
            student_id = str(row[4])[:9]  # Ensure student_id is a string
            present_students.add(student_id)
    
    # Calculate absent students
    absent_students = [student_id for student_id in student_list if str(student_id) not in present_students]
    
    return present_students, absent_students

def upload_to_s3(data, bucket_name, file_name):
    """Upload the data to S3."""
    s3.put_object(Bucket=bucket_name, Key=file_name, Body=json.dumps(data))

def lambda_handler(event, context):
    # Assume the Excel file is stored in S3 and triggered by an S3 event
    bucket_name = "student-excel-files"
    file_name = "test.xlsx"
    
    # Read the Excel file from S3
    response = s3.get_object(Bucket=bucket_name, Key=file_name)
    excel_file = BytesIO(response['Body'].read())

    # Use openpyxl to read the Excel file
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    
    # Assume the questions are in the first row
    questions = [cell.value for cell in sheet[1]]
    
    student_responses = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        student_id = str(row[4])[:9]  # Ensure student_id is a string
        student_responses[student_id] = {questions[i]: row[i] for i in range(2, len(questions)) if row[i] is not None}

    # Calculate attendance (assuming attendance is in the 1st column, adjust index if needed)
    present_students, absent_students = calculate_attendance(sheet, attendance_column_index=0)
    print(f"出席人數: {len(present_students)}")
    
    results = {}
    for target_question in questions[5:]:  # Skip metadata columns
        specific_question_responses = {student_id: responses[target_question] 
                                       for student_id, responses in student_responses.items() 
                                       if target_question in responses}
                                       

        scores_list = analyze_responses(target_question, specific_question_responses) # [json1, json2, ...]
        # if return value is not a list, retry the request
        if not isinstance(scores_list, list):
            scores_list = analyze_responses(target_question, specific_question_responses, start_index=scores_list)
        results[target_question] = scores_list

    upload_to_s3(results, "analysis-results-reports", "student-response-evaluation.json")
    
    return {
        'statusCode': 200,
        'body': json.dumps({
            'analysis_results': results,
            'attendance': {
                'present_students': list(present_students),
                'absent_students': absent_students
            }
        })
    }